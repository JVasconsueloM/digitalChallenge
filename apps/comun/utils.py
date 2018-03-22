import urllib.parse as urlparse
import openpyxl
import re
# from django.urls import reverse
from django.http import QueryDict
from datetime import datetime

from rest_framework.response import Response
from rest_framework import status

from apps.comun.constants import LIMIT, OFFSET
from apps.comun.models import Product, ProductDetails, Brand


class ProcessExcel(object):
    def __init__(self, **kwargs):
        # obtenemos la instancia del excel en memoria
        self.instance_excel = kwargs.get('instance')
        # procesamos la informacion
        self.process()

    def process(self):
        # abrimos el excel
        wb = openpyxl.load_workbook(self.instance_excel)
        # obtenemos la hoja "product"
        sheet = wb.get_sheet_by_name('PRODUCT')
        # solo recorremos desde la posicion 2 por que la posicion 1 suele usarse para los titulos
        # el max row recorrido solo sera hasta la ultima fila modificada en el excel
        products = []
        brands = []
        product_details = []

        max_rows = sheet.max_row + 1
        for row_index in range(2, max_rows):
            id = sheet.cell(row=row_index, column=1).value
            created = sheet.cell(row=row_index, column=2).value
            modified = sheet.cell(row=row_index, column=3).value
            is_active = sheet.cell(row=row_index, column=4).value
            type = sheet.cell(row=row_index, column=5).value
            name = sheet.cell(row=row_index, column=6).value
            description = sheet.cell(row=row_index, column=7).value
            is_variation = sheet.cell(row=row_index, column=8).value
            brand_id = sheet.cell(row=row_index, column=9).value
            code = sheet.cell(row=row_index, column=10).value
            family = sheet.cell(row=row_index, column=11).value
            is_complement = sheet.cell(row=row_index, column=12).value
            is_delete = sheet.cell(row=row_index, column=13).value

            product = Product()
            product.id = id
            product.created = self.datetime(created)
            product.modified = self.datetime(modified)
            product.is_active = self.boolean(is_active)
            product.type = type
            product.name = name
            product.description = description
            product.is_variation = self.boolean(is_variation)
            product.brand_id = brand_id
            product.code = code
            product.family = family
            product.is_complement = self.boolean(is_complement)
            product.is_delete = self.boolean(is_delete)
            products.append(product)
            brands.append(brand_id)

        # solo se realiza esto con el fin de probar la funcion sin obtener algun error de INTEGRIDAD por la inexistencia
        # del brand
        for brand_id in set(brands):
            Brand.objects.get_or_create(id=brand_id, name=brand_id)

        sheet = wb.get_sheet_by_name('PRODUCT_DETAIL')
        max_rows = sheet.max_row + 1
        for row_index in range(2, max_rows):
            id = sheet.cell(row=row_index, column=1).value
            created = sheet.cell(row=row_index, column=2).value
            modified = sheet.cell(row=row_index, column=3).value
            is_active = sheet.cell(row=row_index, column=4).value
            is_visibility = sheet.cell(row=row_index, column=5).value
            price = sheet.cell(row=row_index, column=6).value
            price_offer = sheet.cell(row=row_index, column=7).value
            offer_day_from = sheet.cell(row=row_index, column=8).value
            offer_dat_to = sheet.cell(row=row_index, column=9).value
            quantity = sheet.cell(row=row_index, column=10).value
            sku = sheet.cell(row=row_index, column=11).value
            product_id = sheet.cell(row=row_index, column=12).value

            detail = ProductDetails()
            detail.id = id
            detail.created = self.datetime(created)
            detail.modified = self.datetime(modified)
            detail.is_active = self.boolean(is_active)
            detail.is_visibility = self.boolean(is_visibility)
            detail.price = price
            detail.price_offer = price_offer
            detail.offer_day_from = self.datetime(offer_day_from)
            detail.offer_dat_to = self.datetime(offer_dat_to)
            detail.quantity = quantity
            detail.sku = sku
            detail.product_id = product_id
            product_details.append(detail)

        # creamos los registros con un bulk create para tener mejor performance
        self.products = Product.objects.bulk_create(products)
        self.product_details = ProductDetails.objects.bulk_create(product_details)

    # las siguientes validaciones/formats son simples para proceder con el challenge
    def datetime(self, value):
        try:
            if not value:
                return

            if isinstance(value, datetime):
                return value

            value = re.compile("[.+]").split(value)[0]
            return datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        except:
            return

    def boolean(self, value):
        string = value.upper() if isinstance(value, str) else False
        return string == 'TRUE'


def build_url(*args, **kwargs):
    # path es el uri absoluto
    # name es path name(o url name) de la url a consultar  por ejemplo 'common:api_products'
    # params son los parametros GET
    path = kwargs.pop('path', None)
    # name = kwargs.pop('name', None)
    params = kwargs.pop('params', {})
    # path = reverse(name) if name else path

    if not params:
        return path

    # partimos el uri para poder obtener los GET params
    parts = list(urlparse.urlparse(path))
    # instanciamos QueryDict para poder usar la funcion urlencode
    # tambien podemos usar la url enconde importando urllib.urlencode(diccionario)
    qdict = QueryDict('', mutable=True)
    # combinamos los paramentros GET de la uri con los nuevos parametros GET
    params = {**dict(urlparse.parse_qsl(parts[4])), **params}
    qdict.update(params)
    parts[4] = qdict.urlencode()
    return urlparse.urlunparse(parts)


def paginate(request, queryset, queryset_count):
    # inicializamos data
    data = {}

    # se obtienen los parametros para la paginación manual
    uri = request.build_absolute_uri()
    limit = int(request.GET.get('limit', LIMIT))
    offset = int(request.GET.get('offset', OFFSET))
    prev_offset = offset - limit
    next_offset = offset + limit
    has_prev = offset
    has_next = queryset_count > next_offset

    # obtenemos una lista de diccionarios y aplicamos la "paginacion"
    # especificamos esta excepcion para personalizar el mensage
    try:
        queryset = queryset.values()[offset:next_offset]
    except AssertionError:
        raise ValidationError('limit or offset can not be negative numbers.')

    # asignamos los valores a su respectivo key, previus & next son urls para consultar la siguiente paginacion
    # en caso existan
    data['count'] = queryset_count
    data['previus'] = build_url(path=uri, params={'limit': limit, 'offset': prev_offset}) if has_prev else None
    data['next'] = build_url(path=uri, params={'limit': limit, 'offset': next_offset}) if has_next else None
    data['results'] = queryset
    return data


def valid_filters(Model):
    _f = []
    # obtenemos los campos del modelo, y los lookups que pueden ser usados por cada tipo de campo
    for field in Model._meta.fields:
        _f.append(field.name)
        for lookup in field.class_lookups.keys():
            _f.append('{0}__{1}'.format(field.name, lookup))

    # retornamos una lista de 'params' validos para realizar un filter en un queryset
    return _f


def exception_response(f):
    def inner(*args, **kwargs):
        try:
            return f(*args, **kwargs)

        except (TypeError, ValidationError) as e:
            return response(e.args, status.HTTP_400_BAD_REQUEST)

        except Product.DoesNotExist as e:
            return response(e.args, status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return response(e.args, status.HTTP_500_INTERNAL_SERVER_ERROR)

    return inner


# creamos una excepcion personalizada para usarla posteriormente en el decorador de excepciones
class ValidationError(Exception):
    def __init__(self, message):
        super(ValidationError, self).__init__(message)


# error reponse format
def response(exception, status):
    return Response({'details': exception}, status=status)


def model_setattr(o, d):
    for attr, val in d.items():
        setattr(o, attr, val)
